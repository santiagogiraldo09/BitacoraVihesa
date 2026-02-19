from flask import Flask, request, jsonify, render_template, send_file, redirect,url_for, flash, jsonify
import azure.cognitiveservices.speech as speechsdk
from azure.storage.blob import BlobServiceClient,BlobClient,ContainerClient
from werkzeug.utils import secure_filename
import base64
import io
from io import BytesIO
from PIL import Image
import os
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from flask_cors import CORS
from datetime import datetime
from azure.storage.blob import ContentSettings
from dotenv import load_dotenv
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.lists.list import List
from office365.sharepoint.listitems.listitem import ListItem
import psycopg2
from werkzeug.security import generate_password_hash, check_password_hash
from flask import session
import secrets
from pydub import AudioSegment
import tempfile
import traceback
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
import base64
import uuid
import json
import requests

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Configuración PostgreSQL
POSTGRES_CONFIG = {
    "host": "localhost",
    "database": "Bitacora",
    "user": "postgres",  # Normalmente 'postgres' por defecto
    "password": "Daniel2030#",
    "port": "5432"  # Puerto predeterminado de PostgreSQL
}

SYNCHRO_CONFIG = {
    'client_id': 'service-o5fkAjNrOy3DBriRDwK4aA3Ud',
    'client_secret': 'VTkTyFi36+pUdJ/drZ5chOEhJufMuAZGofF9fzgg/SOUOkrPhPOZERxsq07FpleSZ0bBIRPJVjOua+bR4Exe3Q==',
    'token_url': 'https://ims.bentley.com/connect/token',
    'forms_url': 'https://api.bentley.com/forms',
    'itwin_id': '29d0867b-2158-4b7a-ae03-c63a7661ca58',
    'form_id': 'e4bQKVghekuuA8Y6dmHKWPFDh67WqydKr1vfz4Z0oAs'  # Formulario 1.09-00001
}

# Configura SharePoint (modifica con tus datos)
SHAREPOINT_SITE_URL = "https://iacsas.sharepoint.com/sites/Pruebasproyectossantiago"
LIST_NAME = "Proyectos"  # Nombre de la biblioteca
LIST_NAME_REGISTROS = "RegistrosBitacora"
SHAREPOINT_USER = "santiago.giraldo@iac.com.co"
SHAREPOINT_PASSWORD = "Latumbanuncamuere3"


# Cargar variables de entorno
#load_dotenv('config/settings.env')  # Ruta relativa al archivo .env

app = Flask(__name__,template_folder='templates')
app.secret_key = secrets.token_hex(16)  # Clave secreta para sesiones
#app.secret_key = '78787878tyg8987652vgdfdf3445'
CORS(app)

projects = []

# Conecta con el servicio de Blob Storage de Azure
connection_string = "DefaultEndpointsProtocol=https;AccountName=registrobitacora;AccountKey=ZyHZAOvOBijiOfY3BR3ZEDZsCAHOu3swEPnS+D7AacR2Yr94HS+jBMa2/20sJpZ71decGXYHQxE2+AStBWI/wA==;EndpointSuffix=core.windows.net"
container_name = "registros"


# Inicializa el cliente de BlobServiceClient
blob_service_client = BlobServiceClient.from_connection_string(connection_string)

# ========================================
# OBTENER TOKEN DE BENTLEY
# ========================================
def obtener_token_synchro():
    """Obtiene token de acceso de Bentley IMS"""
    try:
        payload = {
            'grant_type': 'client_credentials',
            'client_id': SYNCHRO_CONFIG['client_id'],
            'client_secret': SYNCHRO_CONFIG['client_secret'],
            'scope': 'itwin-platform'
        }
        
        response = requests.post(SYNCHRO_CONFIG['token_url'], data=payload, timeout=10)
        
        if response.status_code == 200:
            token = response.json().get('access_token')
            print("✅ Token obtenido")
            return token
        else:
            print(f"❌ Error obteniendo token: {response.status_code}")
            return None
            
    except Exception as e:
        print(f"❌ Excepción: {str(e)}")
        return None

# ========================================
# ENVIAR ACTIVIDADES A SYNCHRO
# ========================================
def enviar_actividades_synchro(token, data):
    """Envía todas las actividades al formulario de Synchro"""
    try:
        headers = {
            'Authorization': f'Bearer {token}',
            'Accept': 'application/vnd.bentley.itwin-platform.v2+json',
            'Prefer': 'return=representation',
            'Content-Type': 'application/json'
        }
        
        # 1. Obtener formulario actual
        url_form = f"{SYNCHRO_CONFIG['forms_url']}/{SYNCHRO_CONFIG['form_id']}"
        response = requests.get(url_form, headers=headers, timeout=10)
        
        if response.status_code != 200:
            return {'success': False, 'error': f'No se pudo obtener formulario: {response.status_code}'}
        
        form_actual = response.json().get('form', {})
        props = form_actual.get('properties', {})
        
        # 2. Actualizar propiedades básicas
        props['Codigo Proyecto'] = data['codigo_proyecto']
        props['Contratista'] = data['contratista']
        props['Contrato'] = data['contrato']
        
        # 3. Sección 1: Actividades finalizadas
        actividades_finalizadas = props.get('Actividades finalizadas', [])
        for act in data.get('actividades_finalizadas', []):
            nueva_act = {
                'id': str(uuid.uuid4()),
                '__x00cd__tem': act['item'],
                'Descripci__x00f3__n': act['descripcion'],
                'Observaciones__x0020__actividades__x': act['observaciones']
            }
            actividades_finalizadas.append(nueva_act)
        props['Actividades finalizadas'] = actividades_finalizadas
        
        # 4. Sección 2: Actividades pendientes por culminar
        actividades_pendientes = props.get('Actividades pendientes', [])
        for act in data.get('actividades_pendientes', []):
            nueva_act = {
                'id': str(uuid.uuid4()),
                '__x00cd__tem__x0020__Pendiente': act['item'],
                'Descripci__x00f3__n__x0020__pendient': act['descripcion'],
                'Pendiente__x0020__generado': act.get('pendiente_generado', ''),
                'Observaciones__x0020__pendientes': act['observaciones']
            }
            actividades_pendientes.append(nueva_act)
        props['Actividades pendientes'] = actividades_pendientes
        
        # 5. Sección 3: Actividades pendientes por facturar
        # Nota: Necesitarás el nombre exacto de este campo en Synchro
        # Por ahora lo dejo como ejemplo
        if 'actividades_facturar' in data and data['actividades_facturar']:
            actividades_facturar = props.get('Actividades pendientes por facturar', [])
            for act in data['actividades_facturar']:
                nueva_act = {
                    'id': str(uuid.uuid4()),
                    '__x00cd__tem': act['item'],
                    'Descripci__x00f3__n': act['descripcion'],
                    'Cantidad_contractual': act['cantidad_contractual'],
                    'Cantidad_facturada': act['cantidad_facturada'],
                    'Cantidad_pendiente': act['cantidad_pendiente'],
                    'Observaci__x00f3__n': act['observacion']
                }
                actividades_facturar.append(nueva_act)
            props['Actividades pendientes por facturar'] = actividades_facturar
        
        # 6-8. Secciones de documentación (similar estructura)
        # Agregar según los nombres exactos de los campos en Synchro
        
        # 9. Enviar actualización
        cambios = {'properties': props}
        response_update = requests.patch(url_form, headers=headers, json=cambios, timeout=15)
        
        if response_update.status_code == 200:
            print("✅ Formulario actualizado en Synchro")
            return {'success': True, 'form_id': SYNCHRO_CONFIG['form_id']}
        else:
            error_msg = response_update.text
            print(f"❌ Error actualizando: {error_msg}")
            return {'success': False, 'error': error_msg}
            
    except Exception as e:
        print(f"❌ Excepción: {str(e)}")
        return {'success': False, 'error': str(e)}

# ========================================
# SUBIR ATTACHMENTS A SYNCHRO
# ========================================
def subir_attachments_synchro(token, fotos, videos):
    """Sube fotos y videos como adjuntos al formulario"""
    try:
        headers = {
            'Authorization': f'Bearer {token}',
            'Accept': 'application/vnd.bentley.itwin-platform.v2+json'
        }
        
        url_attachments = f"{SYNCHRO_CONFIG['forms_url']}/{SYNCHRO_CONFIG['form_id']}/attachments"
        
        contador = 0
        
        # Subir fotos
        for i, foto_base64 in enumerate(fotos[:10]):  # Máximo 10 fotos
            try:
                if ',' in foto_base64:
                    foto_base64 = foto_base64.split(',')[1]
                
                foto_bytes = base64.b64decode(foto_base64)
                
                files = {
                    'file': (f'foto_{i+1}.jpg', io.BytesIO(foto_bytes), 'image/jpeg')
                }
                
                data = {
                    'caption': f'Foto {i+1} - Evidencia'
                }
                
                response = requests.post(url_attachments, headers=headers, files=files, data=data, timeout=30)
                
                if response.status_code == 201:
                    contador += 1
                    print(f"✅ Foto {i+1} subida")
                else:
                    print(f"⚠️ Error subiendo foto {i+1}")
                    
            except Exception as e:
                print(f"⚠️ Error procesando foto {i+1}: {str(e)}")
                continue
        
        # Subir videos
        for i, video_base64 in enumerate(videos[:5]):  # Máximo 5 videos
            try:
                if ',' in video_base64:
                    video_base64 = video_base64.split(',')[1]
                
                video_bytes = base64.b64decode(video_base64)
                
                files = {
                    'file': (f'video_{i+1}.webm', io.BytesIO(video_bytes), 'video/webm')
                }
                
                data = {
                    'caption': f'Video {i+1} - Evidencia'
                }
                
                response = requests.post(url_attachments, headers=headers, files=files, data=data, timeout=60)
                
                if response.status_code == 201:
                    contador += 1
                    print(f"✅ Video {i+1} subido")
                else:
                    print(f"⚠️ Error subiendo video {i+1}")
                    
            except Exception as e:
                print(f"⚠️ Error procesando video {i+1}: {str(e)}")
                continue
        
        return contador
        
    except Exception as e:
        print(f"❌ Error en subir_attachments: {str(e)}")
        return 0


@app.route('/guardar-formulario', methods=['POST'])
def guardar_formulario():
    """Recibe datos del frontend y los envía a Synchro"""
    try:
        data = request.json
        print("📥 Datos recibidos del frontend")
        
        # Validar que al menos venga UNA sección con datos
        secciones_con_datos = sum([
            1 if data.get('actividades_finalizadas') else 0,
            1 if data.get('actividades_pendientes') else 0,
            1 if data.get('actividades_facturar') else 0,
            1 if data.get('documentacion_seguridad') else 0,
            1 if data.get('documentacion_ambiental') else 0,
            1 if data.get('documentacion_calidad') else 0
        ])
        
        if secciones_con_datos == 0:
            return jsonify({
                'success': False,
                'error': 'Debes llenar al menos una sección del formulario'
            }), 400
        
        print(f"✅ Validación OK: {secciones_con_datos} sección(es) con datos")
        
        # 1. Obtener token
        token = obtener_token_synchro()
        if not token:
            return jsonify({
                'success': False,
                'error': 'No se pudo obtener token de Synchro'
            }), 500
        
        # 2. Enviar actividades a Synchro
        resultado = enviar_actividades_synchro(token, data)
        if not resultado['success']:
            return jsonify(resultado), 500
        
        # 3. Subir fotos/videos si existen
        fotos = data.get('fotos', [])
        videos = data.get('videos', [])
        attachments_subidos = 0
        
        if fotos or videos:
            attachments_subidos = subir_attachments_synchro(token, fotos, videos)
        
        # 4. Retornar éxito
        return jsonify({
            'success': True,
            'mensaje': 'Registro guardado en Synchro exitosamente',
            'form_id': SYNCHRO_CONFIG['form_id'],
            'attachments_subidos': attachments_subidos,
            'secciones_guardadas': secciones_con_datos
        }), 200
        
    except Exception as e:
        print(f"❌ Error en /guardar-formulario: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


def create_user(nombre, apellido, email, password, cargo, rol, empresa):
    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        cursor = conn.cursor()
        
        hashed_password = generate_password_hash(password)
        
        cursor.execute(
            """INSERT INTO usuario (name, apellido, email, password, cargo, rol, empresa)
               VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING user_id""",
            (nombre, apellido, email, hashed_password, cargo, rol, empresa)
        )
        
        user_id = cursor.fetchone()[0]
        conn.commit()
        return user_id
    except psycopg2.Error as e:
        print(f"Error al crear usuario: {e}")
        return None
    finally:
        if conn:
            conn.close()

def verify_user(email, password):
    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        cursor = conn.cursor()
        
        cursor.execute(
            "SELECT user_id, password FROM usuario WHERE email = %s",
            (email,)
        )
        
        user = cursor.fetchone()
        if user and check_password_hash(user[1], password):
            return user[0]  # Devuelve el ID del usuario
        return None
    except psycopg2.Error as e:
        print(f"Error al verificar usuario: {e}")
        return None
    finally:
        if conn:
            conn.close()


def insert_registro_bitacora(respuestas, id_proyecto, fotos=None, videos=None):
    """
    Inserta un nuevo registro de bitácora, junto con sus fotos y videos asociados
    y sus descripciones, en la base de datos.
    """
    conn = None  # Definimos conn aquí para asegurarnos de que exista en el bloque finally
    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        cursor = conn.cursor()

        # CAMBIO 1: Simplificamos el INSERT principal.
        # - Eliminamos la columna 'foto_base64' que ya es obsoleta.
        # - Cambiamos los nombres de las claves para que coincidan con tu formulario.
        cursor.execute("""
            INSERT INTO registrosbitacoraeqing (
                zona_intervencion, -- Mapeado desde "Tipo de informe"
                items,             -- Mapeado desde "Sede"
                metros_lineales,   -- Mapeado desde "Repuestos utilizados"
                proximas_tareas,   -- Mapeado desde "Repuestos a cotizar"
                id_proyecto
            )
            VALUES (%s, %s, %s, %s, %s)
            RETURNING id_registro
        """, (
            respuestas.get('zona_intervencion'),
            respuestas.get('items'),
            respuestas.get('metros_lineales'),
            respuestas.get('proximas_tareas'),
            id_proyecto,
        ))
        id_registro = cursor.fetchone()[0]

        # CAMBIO 2: Actualizamos el bucle para que maneje objetos (archivo + descripción).
        # Ahora esperamos una lista de diccionarios, no solo una lista de strings.
        for foto_obj in fotos or []:
            file_data = foto_obj.get('file_data')
            description = foto_obj.get('description')
            cursor.execute(
                """INSERT INTO fotos_registro 
                   (id_registro, imagen_base64, description) 
                   VALUES (%s, %s, %s)""",
                (id_registro, file_data, description)
            )

        # CAMBIO 3: Hacemos lo mismo para los videos.
        for video_obj in videos or []:
            file_data = video_obj.get('file_data')
            description = video_obj.get('description')
            cursor.execute(
                """INSERT INTO videos_registro 
                   (id_registro, video_base64, description) 
                   VALUES (%s, %s, %s)""",
                (id_registro, file_data, description)
            )

        conn.commit()
        print(f"Registro {id_registro} guardado exitosamente en PostgreSQL.")

    except psycopg2.Error as e: # MEJORA: Capturamos el error específico de psycopg2 para más detalles
        print(f"Error de base de datos al guardar en PostgreSQL: {e}")
        # Opcional: podrías querer que la función devuelva un error
        # raise e 
    except Exception as e:
        print(f"Error general al guardar en PostgreSQL: {str(e)}")
        # raise e
    finally:
        if conn:
            conn.close()

def create_project(user_id, nombre, fecha_inicio, fecha_fin, director, ubicacion, coordenadas, cliente, numero_proyecto):
    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        #conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute(
            """INSERT INTO proyectos (nombre_proyecto, fecha_inicio, fecha_fin, director_obra, ubicacion, coordenadas, user_id, cliente, numero_proyecto)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id_proyecto""",
            (nombre, fecha_inicio, fecha_fin, director, ubicacion, coordenadas, user_id, cliente, numero_proyecto)
        )
        
        project_id = cursor.fetchone()[0]
        conn.commit()
        return project_id
    except psycopg2.Error as e:
        print(f"Error al crear proyecto: {e}")
        return None
    finally:
        if conn:
            conn.close()

def get_user_projects(user_id):
    conn = None
    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        #conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute(
            """SELECT id_proyecto, nombre_proyecto, fecha_inicio, cliente, user_id 
               FROM "proyectosVihesa" WHERE user_id = %s ORDER BY fecha_inicio DESC""",
            (user_id,)
        )
        
        projects = []
        for row in cursor.fetchall():
            projects.append({
                'id_proyecto': row[0],
                'name': row[1],
                'fecha_inicio': row[2].strftime('%Y-%m-%d'),
                'cliente': row[3],
                'user_id': row[4],

            })
        
        return projects
    except psycopg2.Error as e:
        print(f"Error al obtener proyectos: {e}")
        return []
    finally:
        if conn:
            conn.close()

# Función para subir archivos a Azure Blob Storage
def upload_to_blob(file_name, data, content_type):
    try:
        blob_client = blob_service_client.get_blob_client(container=container_name, blob=file_name)
        blob_client.upload_blob(data, blob_type="BlockBlob", content_settings={"content_type": content_type})
        print(f"Archivo {file_name} subido con éxito.")
    except Exception as e:
        print(f"Error al subir {file_name}: {e}")
        raise


def get_speech_config():
    speech_key = '999fcb4d3f34436ab454ec47920febe0'
    service_region = 'centralus'
    speech_config = speechsdk.SpeechConfig(subscription=speech_key, region=service_region)
    speech_config.speech_recognition_language = "es-CO"
    speech_config.speech_synthesis_language = "es-CO"
    speech_config.speech_synthesis_voice_name = "es-CO-GonzaloNeural"
    speech_config.set_property(speechsdk.PropertyId.SpeechServiceConnection_EndSilenceTimeoutMs, "8000")

    # Esto le pide a Azure que formatee el texto, convirtiendo palabras como "cinco" a "5".
    speech_config.set_property(speechsdk.PropertyId.SpeechServiceResponse_PostProcessingOption, "TrueText")

    return speech_config

def synthesize_speech(text):
    speech_config = get_speech_config()
    audio_config = speechsdk.audio.AudioOutputConfig(use_default_speaker=True)
    synthesizer = speechsdk.SpeechSynthesizer(speech_config=speech_config, audio_config=audio_config)
    result = synthesizer.speak_text_async(text).get()
    return result.reason == speechsdk.ResultReason.SynthesizingAudioCompleted

#Obtener los proyectos desde Azure Blob Storage
def get_projects_from_blob():
    projects = []
    try:
        # Obtener el cliente del contenedor
        container_client = blob_service_client.get_container_client(container_name)
        
        # Listar los blobs en el directorio de proyectos
        blobs = list(container_client.list_blobs(name_starts_with="Proyectos/"))
        
        for blob in blobs:
            if blob.name.endswith('.txt'):
                # Obtener el cliente del blob
                blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob.name)
                
                # Descargar el contenido del blob
                content = blob_client.download_blob().readall().decode('utf-8')
                
                # Extraer información del proyecto
                project_info = {}
                for line in content.strip().split('\n'):
                    line = line.strip()
                    if line:
                        parts = line.split(':', 1)
                        if len(parts) == 2:
                            key = parts[0].strip()
                            value = parts[1].strip()
                            project_info[key] = value
                
                # Extraer el nombre del proyecto del nombre del archivo
                file_name = blob.name.split('/')[-1]
                project_name = file_name.replace('proyecto_', '').replace('.txt', '')
                
                # Crear un objeto de proyecto
                project = {
                    'name': project_info.get('Nombre del Proyecto', project_name),
                    'date': project_info.get('Fecha de Inicio', 'Fecha no disponible'),
                    'blob_name': blob.name,
                    # Añadir más campos según sea necesario
                }
                
                projects.append(project)
                
    except Exception as e:
        print(f"Error al obtener proyectos del Blob Storage: {e}")
    
    return projects

@app.after_request
def add_header(response):
    response.headers["ngrok-skip-browser-warning"] = "true"
    return response

@app.route('/')
def principalscreen():
    return render_template('PrincipalScreen.html')

@app.route('/paginaprincipal')
def paginaprincipal():
    project_id = request.args.get('project_id')
    project_name = request.args.get('project')

    if not project_id:
        return redirect(url_for('history')) # <--- AQUÍ ES DONDE TE ESTÁ MANDANDO

    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        cursor = conn.cursor()
        
        # ASEGÚRATE DE QUE ESTE QUERY USE LA TABLA NUEVA
        cursor.execute('SELECT * FROM "proyectosVihesa" WHERE id_proyecto = %s', (project_id,))
        proyecto = cursor.fetchone()
        
        if not proyecto:
            # Si el ID existe en la URL pero no en la tabla, te manda a history
            return redirect(url_for('history')) 

        return render_template('paginaprincipal.html', project_id=project_id, project_name=project_name)
    except Exception as e:
        print(f"Error: {e}")
        return redirect(url_for('history'))

@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        apellido = request.form.get('apellido')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        empresa = request.form.get('empresa')
        cargo = request.form.get('cargo')
        rol = request.form.get('rol')
        
        if password != confirm_password:
            flash('Las contraseñas no coinciden', 'error')
            return redirect(url_for('registro'))
        
        user_id = create_user(nombre, apellido, email, password, empresa, cargo, rol)
        if user_id:
            flash('Registro exitoso. Por favor inicie sesión.', 'success')
            return redirect(url_for('principalscreen'))
        else:
            flash('Error al registrar el usuario', 'error')
    
    return render_template('registro.html')

@app.route('/login', methods=['POST'])
def login():
    email = request.form.get('email')
    password = request.form.get('password')
    
    # Validación básica de campos vacíos
    if not email or not password:
        flash('Por favor ingrese ambos campos: email y contraseña', 'error')
        return redirect(url_for('principalscreen'))

    user_id = verify_user(email, password)
    if user_id:
        # Aquí puedes implementar sesiones o JWT
        session['user_id'] = user_id #Establecer sesión
        flash('Inicio de sesión exitoso', 'success')
        return redirect(url_for('registros'))
    else:
        flash('Email o contraseña incorrectos', 'error')
        return redirect(url_for('principalscreen'))
"""
@app.route('/index')
def index():
    return render_template('index.html')
"""

@app.route('/index')
def index():
    if 'user_id' not in session:
        return redirect(url_for('principalscreen'))
    
    project_id = request.args.get('project_id')
    project_info = None
    
    if project_id:
        try:
            conn = psycopg2.connect(**POSTGRES_CONFIG)
            cursor = conn.cursor()
            # Consultamos las columnas exactas de tu tabla según las imágenes
            cursor.execute("""
                SELECT nombre_proyecto, cliente, contratista, orden_de_trabajo, planta, ubicacion 
                FROM "proyectosVihesa" 
                WHERE id_proyecto = %s
            """, (project_id,))
            row = cursor.fetchone()
            if row:
                project_info = {
                    'nombre': row[0],
                    'cliente': row[1],
                    'contratista': row[2],
                    'orden_de_trabajo': row[3],
                    'planta': row[4],
                    'ubicacion': row[5]
                }
            conn.close()
        except Exception as e:
            print(f"Error al obtener info del proyecto: {e}")

    return render_template('index.html', project=project_info)

def obtener_token():
    """Obtiene un token de autenticación de Bentley."""
    try:
        payload = {
            'grant_type': 'client_credentials',
            'client_id': SYNCHRO_CONFIG['client_id'],
            'client_secret': SYNCHRO_CONFIG['client_secret'],
            'scope': 'itwin-platform'
        }
        response = requests.post(SYNCHRO_CONFIG['token_url'], data=payload)
        
        if response.status_code == 200:
            return response.json()['access_token']
        else:
            print(f"Error al obtener token (código {response.status_code}): {response.text}")
            return None
    except Exception as e:
        print(f"Excepción al obtener token: {str(e)}")
        return None
    
def obtener_id_por_numero(token, numero):
    """Busca un formulario por su número y retorna su ID y el objeto 'form' completo."""
    headers = {
        'Authorization': f'Bearer {token}',
        'Accept': 'application/vnd.bentley.itwin-platform.v2+json',
        'Prefer': 'return=representation'
    }
    
    url = SYNCHRO_CONFIG['forms_url']
    params = {
        'iTwinId': SYNCHRO_CONFIG['itwin_id'],
        '$top': 50  # Obtener de 50 en 50
    }
    
    while True:
        try:
            response = requests.get(url, headers=headers, params=params)
            
            if response.status_code != 200:
                print(f"Error al buscar formulario (código {response.status_code}): {response.text}")
                return None, None

            data = response.json()
            forms_data = data.get('forms', data) # A veces la respuesta no viene anidada
            
            forms_list = forms_data.get('formDataInstances', [])
            
            for form in forms_list:
                if form.get('number') == numero:
                    # ¡Encontrado! Retorna el ID y el objeto
                    return form.get('id'), form
            
            # Lógica de paginación
            next_link_data = forms_data.get('_links', {}).get('next')
            if not next_link_data:
                break # No hay más páginas
            
            # Extraer el 'continuationToken' para la siguiente página
            next_href = next_link_data.get('href', '')
            if 'continuationToken=' in next_href:
                params['continuationToken'] = next_href.split('continuationToken=')[-1]
                params.pop('$top', None) # Ya no es necesario
            else:
                break # No se pudo encontrar el token de paginación
                
        except Exception as e:
            print(f"Excepción al buscar formulario: {str(e)}")
            return None, None

    # Si sale del bucle sin encontrarlo
    print(f"No se encontró ningún formulario con el número: {numero}")
    return None, None

@app.route('/formulario-synchro')
def formulario_synchro():
    if 'user_id' not in session:
        return redirect(url_for('principalscreen'))
    # Asumiendo que has guardado el archivo como 'indexFormulario.html' en tu carpeta 'templates'
    return render_template('indexFormulario.html')

@app.route('/get-synchro-form-data')
def get_synchro_data():
    form_number = request.args.get('form_number')
    if not form_number:
        return jsonify({'error': 'Falta form_number'}), 400

    token = obtener_token() # (Tu función de Synchro)
    if not token:
        return jsonify({'error': 'No se pudo obtener el token'}), 500

    form_id, form_data = obtener_id_por_numero(token, form_number) # (Tu función de Synchro)

    if not form_id:
        return jsonify({'error': 'Formulario no encontrado'}), 404

    # Devuelve el 'number' y las 'properties'
    return jsonify({
        'id': form_id,
        'number': form_data.get('number'),
        'properties': form_data.get('properties', {})
    })

@app.route('/update-synchro-form', methods=['POST'])
def update_synchro_data():
    try:
        data = request.json
        form_number = data.get('form_number')
        new_properties = data.get('properties')
        # Aquí también puedes manejar data.get('media')

        if not form_number or not new_properties:
            return jsonify({'error': 'Faltan datos (form_number, properties)'}), 400

        token = obtener_token()
        if not token:
            return jsonify({'error': 'No se pudo obtener el token'}), 500

        form_id, form = obtener_id_por_numero(token, form_number)
        if not form_id:
            return jsonify({'error': 'Formulario no encontrado'}), 404

        props_actuales = form.get('properties', {})
        
        for section_name, new_items in new_properties.items():
            if not new_items: 
                continue

            lista_actual = props_actuales.get(section_name, [])
            
            # Generar UUIDs para los nuevos items ANTES de agregarlos
            for item in new_items:
                item['id'] = str(uuid.uuid4()) # Aseguramos un ID único
            
            lista_actual.extend(new_items)
            props_actuales[section_name] = lista_actual

        # --- PREPARAR Y ENVIAR EL PATCH ---
        headers = {
            'Authorization': f'Bearer {token}',
            'Accept': 'application/vnd.bentley.itwin-platform.v2+json',
            'Content-Type': 'application/json'
        }
        
        cambios = {
            'properties': props_actuales
        }
        
        # --- CAMBIO 1: Usar SYNCHRO_CONFIG en lugar de BASE_URL ---
        url = f"{SYNCHRO_CONFIG['forms_url']}/{form_id}"
        
        # --- CAMBIO 2: Usar requests.patch (con 's') ---
        response = requests.patch(url, headers=headers, json=cambios)
        
        if response.status_code == 200:
            return jsonify(response.json())
        else:
            print(f"Error al actualizar Synchro ({response.status_code}): {response.text}")
            return jsonify({'error': 'Error al actualizar Synchro', 'details': response.text}), response.status_code

    except Exception as e:
        print(f"Excepción en update_synchro_data: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error interno del servidor: {str(e)}'}), 500

@app.route('/formulario')
def indexFormulario():
    """Muestra el formulario con datos pre-cargados"""
    proyecto = {
        'codigo': '10111',
        'contratista': 'ABCD',
        'contrato': '001'
    }
    return render_template('indexFormulario.html', proyecto=proyecto)

@app.route('/registros')
def registros():
    if 'user_id' not in session:
        return redirect(url_for('principalscreen'))
    
    # Obtener proyectos de PostgreSQL
    db_projects = get_user_projects(session['user_id'])
    
    # Obtener proyectos de Azure Blob (si aún los necesitas)
    #blob_projects = get_projects_from_blob()  # Tu función existente
    
    # Combinar proyectos (o usar solo los de PostgreSQL)
    return render_template('registros.html', 
                         db_projects=db_projects)

# Ruta para la vista "history"
@app.route('/history')
def history():
    if 'user_id' not in session:
        return redirect(url_for('principalscreen'))
    
    # Obtenemos los proyectos de PostgreSQL
    db_projects = get_user_projects(session['user_id'])
    
    # IMPORTANTE: get_user_projects ya devuelve 'name', 
    # pero asegúrate de que el HTML lo use correctamente.
    return render_template('history.html', db_projects=db_projects)

@app.route('/usuario')
def usuario():
    return render_template('usuario.html')

@app.route('/inventario')
def inventario():
    return render_template('inventario.html')

# En tu archivo app.py

@app.route('/historialRegistro/<int:id_proyecto>')
def historialregistro(id_proyecto):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    conn = None
    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        cursor = conn.cursor()

        # 1. Info del proyecto (Solo esta lleva comillas dobles según tu indicación)
        cursor.execute('SELECT nombre_proyecto, cliente FROM "proyectosVihesa" WHERE id_proyecto = %s', (id_proyecto,))
        proyecto_info = cursor.fetchone()

        # 2. Reportes (Tablas sin comillas)
        cursor.execute("""
            SELECT id_reporte, fecha FROM reporteDeTrabajoVihesa 
            WHERE id_proyecto = %s ORDER BY fecha DESC
        """, (id_proyecto,))
        reportes_rows = cursor.fetchall()

        reportes_completos = []
        for r_row in reportes_rows:
            id_rep, fecha_dt = r_row
            
            # 3. Equipos (Sin comillas)
            cursor.execute('SELECT nombre_equipo FROM reporteEquiposVihesa WHERE id_reporte = %s', (id_rep,))
            equipos = [e[0] for e in cursor.fetchall()]

            # 4. Notas (Sin comillas)
            cursor.execute('SELECT id_nota_item, nota_texto FROM reporteNotasVihesa WHERE id_reporte = %s', (id_rep,))
            notas_rows = cursor.fetchall()
            
            notas_data = []
            for n_row in notas_rows:
                id_nota, texto = n_row
                # 5 y 6. Multimedia (Sin comillas)
                # 5. Fotos
                cursor.execute('SELECT imagen_base64 FROM fotos_registro_vihesa WHERE id_nota_item = %s', (id_nota,))
                fotos_raw = cursor.fetchall()
                fotos = []
                for f in fotos_raw:
                    img_str = f[0]
                    # Si el string contiene el encabezado, lo quitamos para enviarlo limpio
                    if img_str and "," in img_str:
                        img_str = img_str.split(",")[1]
                    fotos.append(img_str)

                # 6. Videos
                cursor.execute('SELECT video_base64 FROM videos_registro_vihesa WHERE id_nota_item = %s', (id_nota,))
                videos_raw = cursor.fetchall()
                videos = []
                for v in videos_raw:
                    vid_str = v[0]
                    if vid_str and "," in vid_str:
                        vid_str = vid_str.split(",")[1]
                    videos.append(vid_str)

                notas_data.append({'texto': texto, 'fotos': fotos, 'videos': videos})

            reportes_completos.append({
                'id_reporte': id_rep,
                'fecha': fecha_dt.strftime('%d/%m/%Y') if fecha_dt else "S/F",
                'equipos': equipos,
                'notas': notas_data
            })

        return render_template('historialRegistro.html', 
                               proyecto=proyecto_info, 
                               reportes=reportes_completos, 
                               id_proyecto=id_proyecto)
    except Exception as e:
        print(f"Error: {e}")
        return redirect(url_for('history'))
    finally:
        if conn: conn.close()

@app.route('/disciplinerecords')
def disciplinerecords():
    return render_template('disciplinerecords.html')

@app.route('/projectdetails')
def projectdetails():
    return render_template('projectdetails.html')

import json

@app.route('/guardar_reporte_vihesa', methods=['POST'])
def guardar_reporte_vihesa():
    data = request.json
    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        cursor = conn.cursor()

        # PASO 1: Tabla Maestra (reporteDeTrabajoVihesa)
        cursor.execute("""
            INSERT INTO reporteDeTrabajoVihesa (id_proyecto, fecha, user_id) 
            VALUES (%s, %s, %s) RETURNING id_reporte
        """, (data['id_proyecto'], data['fecha'], session['user_id']))
        id_reporte = cursor.fetchone()[0]

        # PASO 2: Tabla Equipos (reporteEquiposVihesa)
        for equipo in data.get('equipos', []):
            cursor.execute("""
                INSERT INTO reporteEquiposVihesa (id_reporte, nombre_equipo) 
                VALUES (%s, %s)
            """, (id_reporte, equipo))

        # PASO 3, 4 y 5: Bucle de Notas y Multimedia
        for nota in data.get('notas', []):
            # Guardar Nota (reporteNotasVihesa)
            cursor.execute("""
                INSERT INTO reporteNotasVihesa (id_reporte, nota_texto) 
                VALUES (%s, %s) RETURNING id_nota_item
            """, (id_reporte, nota['texto']))
            id_nota_item = cursor.fetchone()[0]

            # Guardar Fotos (fotos_registro_vihesa) usando el id_nota_item generado
            for foto_b64 in nota.get('fotos', []):
                cursor.execute("""
                    INSERT INTO fotos_registro_vihesa (id_nota_item, imagen_base64, description) 
                    VALUES (%s, %s, %s)
                """, (id_nota_item, foto_b64, "Evidencia Fotográfica"))

            # Guardar Videos (videos_registro_vihesa)
            for video_b64 in nota.get('videos', []):
                cursor.execute("""
                    INSERT INTO videos_registro_vihesa (id_nota_item, video_base64) 
                    VALUES (%s, %s)
                """, (id_nota_item, video_b64))

        conn.commit()
        return jsonify({"status": "success", "message": "Reporte y fotos guardados"}), 201
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({"status": "error", "error": str(e)}), 500
    finally:
        if conn: conn.close()

@app.route('/add_project', methods=['GET', 'POST'])
def add_project():
    if 'user_id' not in session:
        return jsonify({"error": "No autorizado"}), 401
    
    if request.method == 'POST':
        try:
            data = request.json
            conn = psycopg2.connect(**POSTGRES_CONFIG)
            cursor = conn.cursor()
            
            # Query ajustado a la tabla "proyectosVihesa" con sus 10 columnas
            cursor.execute("""
                INSERT INTO "proyectosVihesa" (
                    nombre_proyecto, fecha_inicio, fecha_fin, cliente, contratista, 
                    orden_de_trabajo, planta, ubicacion, user_id
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s) 
                RETURNING id_proyecto
            """, (
                data.get('project-name'), data.get('start-date'), data.get('end-date'),
                data.get('cliente'), data.get('contratista'), data.get('orden-trabajo'),
                data.get('planta'), data.get('location'), session['user_id']
            ))
            
            nuevo_id = cursor.fetchone()[0]
            conn.commit()
            cursor.close()
            conn.close()
            
            return jsonify({"status": "success", "message": f"Proyecto registrado exitosamente"}), 201
            
        except Exception as e:
            print(f"Error en BD: {str(e)}")
            return jsonify({"status": "error", "error": str(e)}), 500

    return render_template('addproject.html')


@app.route('/ask', methods=['POST'])
def ask_question_route():
    data = request.json
    question = data.get('question', '')
    if not question:
        return jsonify({'error': 'No question provided'}), 400

    success = synthesize_speech(question)
    if success:
        return jsonify({'response': ''}), 200
    else:
        return jsonify({'error': 'Error al sintetizar la pregunta.'}), 500


@app.route('/guardar-inspeccion', methods=['POST'])
def guardar_inspeccion():
    try:
        data = request.json
        project_id = data.get('project_id')
        items = data.get('items', [])

        if not project_id or not items:
            return jsonify({'success': False, 'error': 'Datos incompletos'}), 400

        conn = psycopg2.connect(**POSTGRES_CONFIG)
        cursor = conn.cursor()

        for item in items:
            cursor.execute("""
                INSERT INTO reporte_fiscalizacion (
                    id_proyecto, 
                    edificacion_zona, 
                    item_numero, 
                    area_inspeccionada, 
                    especificacion_tecnica, 
                    condicion_observada, 
                    cumple, 
                    observaciones, 
                    acciones_correctivas
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                project_id,
                item['edificacion_zona'],
                item['item_numero'],
                item['area_inspeccionada'],
                item['especificacion_tecnica'],
                item['condicion_observada'],
                item['cumple'],
                item['observaciones'],
                item['acciones_correctivas']
            ))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'mensaje': 'Inspección guardada correctamente'})
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/guardar-registro', methods=['POST'])
def guardar_registro():
    conn = None
    try:
        data = request.get_json()
        project_id = data.get('project_id')
        items = data.get('items', [])
        # Nota: 'fotos' y 'videos' ahora deberían venir dentro de cada objeto en 'items'
        
        if not project_id or not items:
            return jsonify({"error": "Faltan datos requeridos."}), 400

        conn = psycopg2.connect(**POSTGRES_CONFIG)
        cursor = conn.cursor()

        # 1. Bucle principal para guardar cada ítem
        for item in items:
            cursor.execute("""
                INSERT INTO reporte_fiscalizacion (
                    id_proyecto, edificacion_zona, item_numero, area_inspeccionada, 
                    especificacion_tecnica, condicion_observada, cumple, 
                    observaciones, acciones_correctivas
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id_reporte
            """, (
                project_id, item.get('edificacion_zona'), item.get('item_numero'),
                item.get('area_inspeccionada'), item.get('especificacion_tecnica'),
                item.get('condicion_observada'), item.get('cumple'),
                item.get('observaciones'), item.get('acciones_correctivas')
            ))
            
            # Capturamos el ID específico de ESTE ítem recién insertado
            id_item_actual = cursor.fetchone()[0]

            # 2. GUARDAR FOTOS ESPECÍFICAS DE ESTE ÍTEM (NUEVA UBICACIÓN)
            # El frontend ahora debe enviar las fotos dentro de cada item
            fotos_item = item.get('fotos', []) 
            for foto_obj in fotos_item:
                cursor.execute("""
                    INSERT INTO fotos_registro (id_registro, imagen_base64, description) 
                    VALUES (%s, %s, %s)
                """, (id_item_actual, foto_obj.get('file_data'), foto_obj.get('description')))

            # 3. GUARDAR VIDEOS ESPECÍFICOS DE ESTE ÍTEM
            videos_item = item.get('videos', [])
            for video_obj in videos_item:
                cursor.execute("""
                    INSERT INTO videos_registro (id_registro, video_base64, description) 
                    VALUES (%s, %s, %s)
                """, (id_item_actual, video_obj.get('file_data'), video_obj.get('description')))

        conn.commit()
        return jsonify({"mensaje": "¡Reporte guardado!"}), 200

    except Exception as e:
        if conn: conn.rollback()
        print(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn: conn.close()

@app.route('/eliminar-proyecto', methods=['POST'])
def eliminar_proyecto():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    data = request.get_json()
    proyecto_id = data.get('id_proyecto')

    if not proyecto_id:
        return jsonify({'error': 'Falta el ID del proyecto'}), 400

    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        cursor = conn.cursor()

        # Asegurarse de que el proyecto pertenece al usuario
        cursor.execute("""
            DELETE FROM proyectos
            WHERE id_proyecto = %s AND user_id = %s
        """, (proyecto_id, session['user_id']))
        conn.commit()

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route('/transcribe-audio', methods=['POST'])
def transcribe_audio():
    try:
        if 'audio' not in request.files:
            print("🔴 No se recibió archivo de audio.")
            return jsonify({"error": "No se envió el archivo de audio"}), 400

        file = request.files['audio']
        print(f"📥 Recibido archivo: {file.filename}")

        # Guardar el archivo temporalmente
        temp_input = tempfile.NamedTemporaryFile(delete=False, suffix=".webm")
        file.save(temp_input.name)
        print(f"💾 Guardado en: {temp_input.name}")

        temp_wav = tempfile.NamedTemporaryFile(delete=False, suffix=".wav")
        formato_detectado = None

        try:
            print("🔍 Intentando decodificar como webm...")
            audio = AudioSegment.from_file(temp_input.name, format="webm")
            print("✅ Decodificado como webm.")
            formato_detectado = "webm"
        except Exception as e_webm:
            print("⚠️ Falla al decodificar como webm:", str(e_webm))
            try:
                print("🔁 Intentando decodificar como mp4...")
                audio = AudioSegment.from_file(temp_input.name, format="mp4")
                print("✅ Decodificado como mp4.")
                formato_detectado = "mp4"
            except Exception as e_mp4:
                print("❌ Fallo total al decodificar audio.")
                traceback.print_exc()
                return jsonify({
                    "error": "No se pudo procesar el audio.",
                    "error_webm": str(e_webm),
                    "error_mp4": str(e_mp4)
                }), 500

        # Exportar a WAV
        audio.export(temp_wav.name, format="wav")
        print("🔄 Exportado a WAV:", temp_wav.name)

        # Transcribir con Azure
        speech_config = get_speech_config()
        audio_config = speechsdk.audio.AudioConfig(filename=temp_wav.name)
        recognizer = speechsdk.SpeechRecognizer(speech_config=speech_config, audio_config=audio_config)
        result = recognizer.recognize_once_async().get()

        if result.reason == speechsdk.ResultReason.RecognizedSpeech:
            print("✅ Texto reconocido:", result.text)
            return jsonify({
                "text": result.text,
                "formato_detectado": formato_detectado
            })
        else:
            print("⚠️ No se reconoció el audio:", result.reason)
            return jsonify({
                "error": "No se reconoció el audio.",
                "formato_detectado": formato_detectado
            }), 400

    except Exception as e:
        print("❌ Error general en transcribe_audio:", str(e))
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

#Exportar registros seleccionados a Excel
@app.route('/exportar-registros-excel', methods=['POST'])
def exportar_registros_excel():
    registro_ids = request.form.getlist('registro_ids')
    project_id = request.form.get('project_id')

    if not registro_ids and not project_id:
        return "No se seleccionaron registros ni proyecto", 400

    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        cursor = conn.cursor()

        if not registro_ids:
            cursor.execute("""
                SELECT id_registro, zona_intervencion, items, metros_lineales, proximas_tareas, foto_base64
                FROM registrosbitacoraeqing
                WHERE id_proyecto = %s
                ORDER BY id_registro DESC
            """, (project_id,))
        else:
            format_ids = tuple(map(int, registro_ids))
            cursor.execute("""
                SELECT id_registro, zona_intervencion, items, metros_lineales, proximas_tareas, foto_base64
                FROM registrosbitacoraeqing
                WHERE id_registro IN %s
                ORDER BY id_registro DESC
            """, (format_ids,))

        rows = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Registros"

        # Encabezado
        ws.append(["ID", "Zona de Intervención", "Ítems", "Metros Lineales", "Próximas Tareas", "Foto"])

        row_index = 2  # Comienza después del encabezado

        for row in rows:
            id_registro, zona, items, metros, tareas, foto_base64 = row
            ws.append([id_registro, zona, items, metros, tareas, ""])  # celda para imagen

            if foto_base64:
                try:
                    header, base64_data = foto_base64.split(',', 1) if ',' in foto_base64 else ('', foto_base64)
                    image_data = base64.b64decode(base64_data)
                    img = Image.open(io.BytesIO(image_data))
                    img.thumbnail((120, 120))  # redimensiona para celda
                    image_io = io.BytesIO()
                    img.save(image_io, format='PNG')
                    image_io.seek(0)

                    img_excel = ExcelImage(image_io)
                    img_excel.anchor = f"F{row_index}"
                    ws.add_image(img_excel)

                    # Ajustar altura de fila
                    ws.row_dimensions[row_index].height = 90
                except Exception as img_err:
                    print(f"Error al procesar imagen para registro {id_registro}: {img_err}")

            row_index += 1

        # Ajuste de anchos de columnas
        ws.column_dimensions['A'].width = 12  # ID
        ws.column_dimensions['B'].width = 30  # Zona de intervención
        ws.column_dimensions['C'].width = 25  # Ítems
        ws.column_dimensions['D'].width = 20  # Metros lineales
        ws.column_dimensions['E'].width = 35  # Próximas tareas
        ws.column_dimensions['F'].width = 18  # Imagen

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output,
                         download_name="registros_bitacora.xlsx",
                         as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        print(f"Error al exportar: {e}")
        return "Error al exportar", 500
    finally:
        if conn:
            conn.close()


@app.route('/exportar-proyectos-excel', methods=['POST'])
def exportar_proyectos_excel():
    project_ids = request.form.getlist('project_ids')
    
    if not project_ids:
        return "No se seleccionaron proyectos", 400

    try:
        conn = psycopg2.connect(**POSTGRES_CONFIG)
        cursor = conn.cursor()
        wb = Workbook()
        wb.remove(wb.active)  # Eliminar hoja por defecto

        for pid in project_ids:
            try:
                pid_int = int(pid)
            except:
                continue

            # Obtener info del proyecto
            cursor.execute("""
                SELECT nombre_proyecto, fecha_inicio, fecha_fin, director_obra, ubicacion, coordenadas
                FROM proyectos WHERE id_proyecto = %s
            """, (pid_int,))
            proyecto = cursor.fetchone()
            if not proyecto:
                continue

            nombre, fecha_inicio, fecha_fin, director, ubicacion, coordenadas = proyecto
            sheet_title = (nombre[:30] or f"Proyecto {pid_int}").strip()
            ws = wb.create_sheet(title=sheet_title)

            # Encabezado de proyecto
            ws.append(["Nombre del Proyecto:", nombre])
            ws.append(["Fecha de Inicio:", str(fecha_inicio)])
            ws.append(["Fecha de Finalización:", str(fecha_fin)])
            ws.append(["Director del Proyecto:", director])
            ws.append(["Ubicación:", ubicacion])
            ws.append(["Coordenadas:", coordenadas])
            ws.append([])

            # Encabezado de registros
            ws.append(["ID", "Zona de Intervención", "Ítems Instalados", "Metros Lineales", "Próximas Tareas", "Foto"])

            # Obtener registros
            cursor.execute("""
                SELECT id_registro, zona_intervencion, items, metros_lineales, proximas_tareas, foto_base64
                FROM registrosbitacoraeqing
                WHERE id_proyecto = %s
                ORDER BY id_registro DESC
            """, (pid_int,))
            registros = cursor.fetchall()

            row_index = 9
            for registro in registros:
                idr, zona, items, metros, tareas, foto = registro
                ws.append([idr, zona, items, metros, tareas, ""])

                if foto:
                    try:
                        header, base64_data = foto.split(',', 1) if ',' in foto else ('', foto)
                        img_data = base64.b64decode(base64_data)
                        img = Image.open(io.BytesIO(img_data))
                        img.thumbnail((120, 120))
                        img_io = io.BytesIO()
                        img.save(img_io, format='PNG')
                        img_io.seek(0)

                        img_excel = ExcelImage(img_io)
                        img_excel.anchor = f"F{row_index}"
                        ws.add_image(img_excel)

                        ws.row_dimensions[row_index].height = 90
                    except Exception as e:
                        print(f"Error en imagen de registro {idr}: {e}")
                row_index += 1

            # Ajustes de columnas
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 35
            ws.column_dimensions['F'].width = 18

        # Generar archivo
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output,
                         download_name="proyectos_exportados.xlsx",
                         as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(f"Error exportando proyectos: {e}")
        return "Error interno al exportar", 500
    finally:
        if conn:
            conn.close()

if __name__ == '__main__':
    app.run(debug=True)